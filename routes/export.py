"""
Export-Routes – Excel & PDF Export für Reporting.
Endpunkte: /reporting/export/excel  /reporting/export/pdf
"""

import io
import os
from datetime import datetime, timedelta
from flask import Blueprint, request, send_file, current_app
from flask_login import login_required, current_user
from models import Campaign

export_bp = Blueprint("export", __name__, url_prefix="/reporting/export")


def _get_campaigns(platform, period):
    """Gemeinsame Kampagnen-Abfrage für beide Export-Typen."""
    mandant = current_user.mandant
    if period > 0:
        since = datetime.utcnow() - timedelta(days=period)
    else:
        since = None

    if current_user.is_admin:
        query = Campaign.query.filter_by(mandant_id=mandant.id)
    else:
        query = Campaign.query.filter_by(niederlassung_id=current_user.niederlassung_id)

    if since:
        query = query.filter(Campaign.created_at >= since)
    if platform != "all":
        query = query.filter(
            (Campaign.platform == platform) | (Campaign.platform == "both")
        )
    return query.order_by(Campaign.created_at.desc()).all()


def _build_kst_rows(campaigns):
    """Baut KST-Aggregation – identisch zu reporting.py."""
    kst_map = {}
    for c in campaigns:
        kst = c.kostenstelle
        if kst not in kst_map:
            kst_map[kst] = {
                "kostenstelle":   kst,
                "name":           c.niederlassung.name if c.niederlassung else "–",
                "campaign_count": 0,
                "spend":          0.0,
                "budget":         0.0,
                "clicks":         0,
                "conversions":    0,
                "impressions":    0,
            }
        kst_map[kst]["campaign_count"] += 1
        kst_map[kst]["spend"]       += c.total_cost or 0
        kst_map[kst]["budget"]      += c.total_budget or 0
        kst_map[kst]["clicks"]      += c.total_clicks or 0
        kst_map[kst]["conversions"] += c.conversions or 0
        kst_map[kst]["impressions"] += c.total_impressions or 0
    return sorted(kst_map.values(), key=lambda x: x["spend"], reverse=True)


# ─────────────────────────────────────────────────────────────
#  EXCEL EXPORT
# ─────────────────────────────────────────────────────────────
@export_bp.route("/excel")
@login_required
def excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                      Border, Side, numbers)
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference
    except ImportError:
        return "openpyxl nicht installiert.", 500

    platform = request.args.get("platform", "all")
    period   = int(request.args.get("period", 30))
    campaigns = _get_campaigns(platform, period)
    kst_rows  = _build_kst_rows(campaigns)

    wb = Workbook()

    # ── Farben & Stile ────────────────────────────────────────
    BLUE      = "003087"
    BLUE_LIGHT= "E8EEF7"
    GRAY      = "6B7280"
    GRAY_LIGHT= "F4F6F9"
    WHITE     = "FFFFFF"
    GREEN     = "16A34A"
    ORANGE    = "D97706"
    RED       = "DC2626"

    header_font   = Font(name="Calibri", bold=True, color=WHITE, size=11)
    header_fill   = PatternFill("solid", fgColor=BLUE)
    sub_fill      = PatternFill("solid", fgColor=BLUE_LIGHT)
    sub_font      = Font(name="Calibri", bold=True, color=BLUE, size=10)
    total_fill    = PatternFill("solid", fgColor=GRAY_LIGHT)
    total_font    = Font(name="Calibri", bold=True, size=10)
    normal_font   = Font(name="Calibri", size=10)
    center_align  = Alignment(horizontal="center", vertical="center")
    left_align    = Alignment(horizontal="left",   vertical="center")
    right_align   = Alignment(horizontal="right",  vertical="center")

    thin_border = Border(
        left  =Side(style="thin", color="E5E7EB"),
        right =Side(style="thin", color="E5E7EB"),
        top   =Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    def style_header_row(ws, row, cols):
        for col in range(1, cols + 1):
            c = ws.cell(row=row, column=col)
            c.font      = header_font
            c.fill      = header_fill
            c.alignment = center_align
            c.border    = thin_border

    def style_data_row(ws, row, cols, even=False):
        fill = PatternFill("solid", fgColor="F9FAFB") if even else None
        for col in range(1, cols + 1):
            c = ws.cell(row=row, column=col)
            c.font   = normal_font
            c.border = thin_border
            if fill:
                c.fill = fill

    # ── Sheet 1: Zusammenfassung ─────────────────────────────
    ws1 = wb.active
    ws1.title = "Zusammenfassung"
    ws1.sheet_view.showGridLines = False

    # Titel-Banner
    ws1.merge_cells("A1:I2")
    title_cell = ws1["A1"]
    title_cell.value     = f"🎯 Adynex – Reporting Export"
    title_cell.font      = Font(name="Calibri", bold=True, size=16, color=WHITE)
    title_cell.fill      = PatternFill("solid", fgColor=BLUE)
    title_cell.alignment = center_align

    # Meta-Info
    ws1.merge_cells("A3:I3")
    meta = ws1["A3"]
    period_label = {0: "Gesamt", 30: "Letzte 30 Tage", 90: "Letzte 90 Tage", 365: "Letztes Jahr"}.get(period, f"{period} Tage")
    platform_label = {"all": "Alle Plattformen", "google": "Google Ads", "microsoft": "Microsoft Ads"}.get(platform, platform)
    meta.value     = f"Mandant: {current_user.mandant.name}  |  Zeitraum: {period_label}  |  Plattform: {platform_label}  |  Erstellt: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    meta.font      = Font(name="Calibri", size=10, color=GRAY)
    meta.fill      = PatternFill("solid", fgColor=BLUE_LIGHT)
    meta.alignment = center_align
    ws1.row_dimensions[3].height = 18

    ws1.row_dimensions[4].height = 8  # Abstand

    # KPI-Summary-Kacheln (Row 5-7)
    total_spend    = sum(c.total_cost or 0 for c in campaigns)
    total_budget   = sum(c.total_budget or 0 for c in campaigns)
    total_conv     = sum(c.conversions or 0 for c in campaigns)
    total_clicks   = sum(c.total_clicks or 0 for c in campaigns)
    total_impr     = sum(c.total_impressions or 0 for c in campaigns)
    active_count   = sum(1 for c in campaigns if c.status == "active")
    cpa            = (total_spend / total_conv) if total_conv > 0 else 0
    ctr            = (total_clicks / total_impr * 100) if total_impr > 0 else 0

    kpi_titles  = ["AKTIVE KAMPAGNEN", "GESAMTAUSGABEN", "CONVERSIONS", "CPA", "CTR"]
    kpi_values  = [
        str(active_count),
        f"{total_spend:,.2f} €",
        str(total_conv),
        f"{cpa:,.2f} €" if cpa else "–",
        f"{ctr:.2f}%" if ctr else "–",
    ]
    kpi_cols = [("A","B"), ("C","D"), ("E","F"), ("G","H"), ("I","I")]

    for i, ((start, end), title, val) in enumerate(zip(kpi_cols, kpi_titles, kpi_values)):
        merge_range = f"{start}5:{end}5" if start != end else f"{start}5"
        if start != end:
            ws1.merge_cells(merge_range)
        lbl = ws1[f"{start}5"]
        lbl.value     = title
        lbl.font      = Font(name="Calibri", bold=True, size=8, color=GRAY)
        lbl.fill      = sub_fill
        lbl.alignment = center_align

        merge_range2 = f"{start}6:{end}6" if start != end else f"{start}6"
        if start != end:
            ws1.merge_cells(merge_range2)
        v = ws1[f"{start}6"]
        v.value     = val
        v.font      = Font(name="Calibri", bold=True, size=14, color=BLUE)
        v.fill      = sub_fill
        v.alignment = center_align

    ws1.row_dimensions[7].height = 8  # Abstand

    # KST-Tabelle Header (Row 8)
    ws1.merge_cells("A8:I8")
    section = ws1["A8"]
    section.value     = "Auswertung nach Kostenstelle"
    section.font      = Font(name="Calibri", bold=True, size=11, color=WHITE)
    section.fill      = PatternFill("solid", fgColor=BLUE)
    section.alignment = left_align

    kst_headers = ["Niederlassung", "KST", "Kampagnen", "Ausgaben (€)", "Budget (€)", "Auslastung", "Klicks", "Conv.", "CPA (€)"]
    for col, h in enumerate(kst_headers, 1):
        c = ws1.cell(row=9, column=col, value=h)
        c.font      = sub_font
        c.fill      = sub_fill
        c.alignment = center_align
        c.border    = thin_border
    ws1.row_dimensions[9].height = 22

    for i, row in enumerate(kst_rows):
        r = 10 + i
        pct = (row["spend"] / row["budget"] * 100) if row["budget"] > 0 else 0
        cpa_val = (row["spend"] / row["conversions"]) if row["conversions"] > 0 else None
        row_data = [
            row["name"],
            f"KST-{row['kostenstelle']}",
            row["campaign_count"],
            row["spend"],
            row["budget"],
            f"{pct:.0f}%",
            row["clicks"],
            row["conversions"],
            f"{cpa_val:,.2f}" if cpa_val else "–",
        ]
        for col, val in enumerate(row_data, 1):
            c = ws1.cell(row=r, column=col, value=val)
            c.font      = normal_font
            c.alignment = right_align if col in (3, 4, 5, 7, 8, 9) else left_align
            c.border    = thin_border
        style_data_row(ws1, r, 9, even=(i % 2 == 1))
        ws1.row_dimensions[r].height = 20

    # Summenzeile
    total_row = 10 + len(kst_rows)
    total_data = ["Gesamt", "", len(campaigns), total_spend, total_budget, "", total_clicks, total_conv,
                  f"{cpa:,.2f}" if cpa else "–"]
    for col, val in enumerate(total_data, 1):
        c = ws1.cell(row=total_row, column=col, value=val)
        c.font      = total_font
        c.fill      = total_fill
        c.alignment = right_align if col in (3, 4, 5, 7, 8, 9) else left_align
        c.border    = thin_border
    ws1.row_dimensions[total_row].height = 22

    # Spaltenbreiten
    col_widths = [22, 10, 12, 14, 14, 12, 10, 8, 12]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Kampagnen-Detail ────────────────────────────
    ws2 = wb.create_sheet("Kampagnen")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:J1")
    t = ws2["A1"]
    t.value     = "Kampagnen-Übersicht"
    t.font      = Font(name="Calibri", bold=True, size=13, color=WHITE)
    t.fill      = PatternFill("solid", fgColor=BLUE)
    t.alignment = center_align
    ws2.row_dimensions[1].height = 28

    camp_headers = ["Kampagne", "KST", "Plattform", "Status", "Ausgaben (€)", "Budget (€)", "Klicks", "Conv.", "CPA (€)", "Erstellt"]
    for col, h in enumerate(camp_headers, 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.font      = sub_font
        c.fill      = sub_fill
        c.alignment = center_align
        c.border    = thin_border
    ws2.row_dimensions[2].height = 22

    STATUS_DE = {
        "active": "Aktiv", "pending_approval": "Warte auf Genehmigung",
        "paused": "Pausiert", "completed": "Abgeschlossen", "rejected": "Abgelehnt",
    }
    PLATFORM_DE = {"google": "Google", "microsoft": "Microsoft", "both": "Google + Microsoft"}

    for i, c in enumerate(campaigns):
        r = 3 + i
        cpa_c = (c.total_cost / c.conversions) if (c.conversions and c.total_cost) else None
        row_data = [
            c.name,
            f"KST-{c.kostenstelle}",
            PLATFORM_DE.get(c.platform, c.platform),
            STATUS_DE.get(c.status, c.status),
            c.total_cost or 0,
            c.total_budget or 0,
            c.total_clicks or 0,
            c.conversions or 0,
            f"{cpa_c:,.2f}" if cpa_c else "–",
            c.created_at.strftime("%d.%m.%Y") if c.created_at else "–",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=col, value=val)
            cell.font      = normal_font
            cell.alignment = right_align if col in (5, 6, 7, 8, 9) else left_align
            cell.border    = thin_border
        if i % 2 == 1:
            for col in range(1, 11):
                ws2.cell(row=r, column=col).fill = PatternFill("solid", fgColor="F9FAFB")
        ws2.row_dimensions[r].height = 20

    camp_widths = [40, 10, 20, 22, 13, 13, 9, 8, 10, 12]
    for i, w in enumerate(camp_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Datei ausgeben ────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Adynex_Reporting_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ─────────────────────────────────────────────────────────────
#  PDF EXPORT
# ─────────────────────────────────────────────────────────────
@export_bp.route("/pdf")
@login_required
def pdf():
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm, cm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                         Table, TableStyle, Image, HRFlowable)
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.patches as mpatches
        import numpy as np
    except ImportError as e:
        return f"Export-Paket fehlt: {e}", 500

    platform = request.args.get("platform", "all")
    period   = int(request.args.get("period", 30))
    campaigns = _get_campaigns(platform, period)
    kst_rows  = _build_kst_rows(campaigns)

    # ── Farbpalette ───────────────────────────────────────────
    BLUE       = colors.HexColor("#003087")
    BLUE_LIGHT = colors.HexColor("#E8EEF7")
    BLUE_MID   = colors.HexColor("#1E4D8C")
    GRAY       = colors.HexColor("#6B7280")
    GRAY_LIGHT = colors.HexColor("#F4F6F9")
    WHITE      = colors.white
    GREEN      = colors.HexColor("#16A34A")
    ORANGE     = colors.HexColor("#D97706")
    RED        = colors.HexColor("#DC2626")

    # Werte für Charts
    total_spend  = sum(c.total_cost or 0 for c in campaigns)
    total_budget = sum(c.total_budget or 0 for c in campaigns)
    total_conv   = sum(c.conversions or 0 for c in campaigns)
    total_clicks = sum(c.total_clicks or 0 for c in campaigns)
    total_impr   = sum(c.total_impressions or 0 for c in campaigns)
    active_count = sum(1 for c in campaigns if c.status == "active")
    cpa          = (total_spend / total_conv) if total_conv > 0 else 0
    ctr          = (total_clicks / total_impr * 100) if total_impr > 0 else 0

    period_label   = {0: "Gesamt", 30: "Letzte 30 Tage", 90: "Letzte 90 Tage", 365: "Letztes Jahr"}.get(period, f"{period} Tage")
    platform_label = {"all": "Alle Plattformen", "google": "Google Ads", "microsoft": "Microsoft Ads"}.get(platform, platform)

    # ── Charts mit Matplotlib ─────────────────────────────────
    chart_paths = []
    plt.rcParams.update({
        "font.family": "sans-serif",
        "font.size": 9,
        "axes.spines.top": False,
        "axes.spines.right": False,
    })

    if kst_rows:
        labels = [f"KST-{r['kostenstelle']}\n{r['name'][:12]}" for r in kst_rows[:8]]
        spends  = [r["spend"]  for r in kst_rows[:8]]
        budgets = [r["budget"] for r in kst_rows[:8]]
        x = range(len(labels))

        # Chart 1: Ausgaben vs. Budget pro KST
        fig1, ax1 = plt.subplots(figsize=(7, 3.2))
        bar_w = 0.38
        bars1 = ax1.bar([i - bar_w/2 for i in x], spends,  bar_w, label="Ausgaben",  color="#003087")
        bars2 = ax1.bar([i + bar_w/2 for i in x], budgets, bar_w, label="Budget",    color="#93C5FD", alpha=0.85)
        ax1.set_xticks(list(x))
        ax1.set_xticklabels(labels, fontsize=7.5)
        ax1.set_ylabel("Euro (€)", fontsize=8)
        ax1.set_title("Ausgaben vs. Budget nach Kostenstelle", fontsize=10, fontweight="bold", color="#003087")
        ax1.legend(fontsize=8)
        ax1.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:,.0f} €"))
        ax1.set_facecolor("#FAFAFA")
        fig1.patch.set_facecolor("#FAFAFA")
        fig1.tight_layout()
        path1 = "/tmp/chart_budget.png"
        fig1.savefig(path1, dpi=150, bbox_inches="tight", facecolor="#FAFAFA")
        plt.close(fig1)
        chart_paths.append(path1)

        # Chart 2: Conversions pro KST
        convs = [r["conversions"] for r in kst_rows[:8]]
        fig2, ax2 = plt.subplots(figsize=(7, 3.0))
        bar_colors = ["#003087" if cv > 0 else "#D1D5DB" for cv in convs]
        ax2.bar(labels, convs, color=bar_colors, width=0.55)
        ax2.set_ylabel("Conversions", fontsize=8)
        ax2.set_title("Conversions nach Kostenstelle", fontsize=10, fontweight="bold", color="#003087")
        ax2.set_facecolor("#FAFAFA")
        fig2.patch.set_facecolor("#FAFAFA")
        fig2.tight_layout()
        path2 = "/tmp/chart_conv.png"
        fig2.savefig(path2, dpi=150, bbox_inches="tight", facecolor="#FAFAFA")
        plt.close(fig2)
        chart_paths.append(path2)

        # Chart 3: CPA pro KST (Balken horizontal)
        cpas = [(r["spend"] / r["conversions"]) if r["conversions"] > 0 else 0 for r in kst_rows[:8]]
        kst_labels_short = [f"KST-{r['kostenstelle']} {r['name'][:10]}" for r in kst_rows[:8]]
        fig3, ax3 = plt.subplots(figsize=(7, 3.0))
        bar_col3 = ["#DC2626" if v > 50 else "#D97706" if v > 20 else "#16A34A" for v in cpas]
        ax3.barh(kst_labels_short, cpas, color=bar_col3, height=0.55)
        ax3.set_xlabel("CPA (€)", fontsize=8)
        ax3.set_title("Cost per Application (CPA) nach Kostenstelle", fontsize=10, fontweight="bold", color="#003087")
        ax3.xaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:,.0f} €"))
        ax3.set_facecolor("#FAFAFA")
        fig3.patch.set_facecolor("#FAFAFA")
        fig3.tight_layout()
        path3 = "/tmp/chart_cpa.png"
        fig3.savefig(path3, dpi=150, bbox_inches="tight", facecolor="#FAFAFA")
        plt.close(fig3)
        chart_paths.append(path3)

    # ── PDF Aufbau ─────────────────────────────────────────────
    output  = io.BytesIO()
    doc     = SimpleDocTemplate(
        output,
        pagesize=A4,
        leftMargin=18*mm, rightMargin=18*mm,
        topMargin=15*mm,  bottomMargin=15*mm,
        title="Adynex Reporting",
        author="Adynex SaaS",
    )

    styles = getSampleStyleSheet()
    style_h1     = ParagraphStyle("H1",     parent=styles["Normal"], fontSize=20, textColor=WHITE,  fontName="Helvetica-Bold",  spaceAfter=0, leading=26)
    style_h2     = ParagraphStyle("H2",     parent=styles["Normal"], fontSize=13, textColor=BLUE,   fontName="Helvetica-Bold",  spaceAfter=4, spaceBefore=12)
    style_body   = ParagraphStyle("Body",   parent=styles["Normal"], fontSize=9,  textColor=GRAY,   spaceAfter=4)
    style_center = ParagraphStyle("Center", parent=styles["Normal"], fontSize=9,  textColor=GRAY,   alignment=TA_CENTER)
    style_kpi_lbl= ParagraphStyle("KpiLbl", parent=styles["Normal"], fontSize=7.5,textColor=GRAY,   fontName="Helvetica-Bold",  alignment=TA_CENTER, spaceAfter=2)
    style_kpi_val= ParagraphStyle("KpiVal", parent=styles["Normal"], fontSize=18, textColor=BLUE,   fontName="Helvetica-Bold",  alignment=TA_CENTER)
    style_kpi_sub= ParagraphStyle("KpiSub", parent=styles["Normal"], fontSize=8,  textColor=GRAY,   alignment=TA_CENTER)

    story = []

    # ── Header-Banner ──────────────────────────────────────────
    header_data = [[Paragraph("🎯 Adynex – Reporting & Analytics", style_h1)]]
    header_tbl  = Table(header_data, colWidths=[174*mm])
    header_tbl.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,-1), BLUE),
        ("TOPPADDING",  (0,0), (-1,-1), 12),
        ("BOTTOMPADDING",(0,0),(-1,-1), 12),
        ("LEFTPADDING", (0,0), (-1,-1), 14),
        ("ROWBACKGROUNDS",(0,0),(-1,-1),[BLUE]),
    ]))
    story.append(header_tbl)
    story.append(Spacer(1, 4*mm))

    # Meta-Zeile
    story.append(Paragraph(
        f"<b>Mandant:</b> {current_user.mandant.name} &nbsp;|&nbsp; "
        f"<b>Zeitraum:</b> {period_label} &nbsp;|&nbsp; "
        f"<b>Plattform:</b> {platform_label} &nbsp;|&nbsp; "
        f"<b>Erstellt:</b> {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        style_body
    ))
    story.append(Spacer(1, 5*mm))

    # ── KPI-Kacheln ────────────────────────────────────────────
    story.append(Paragraph("Gesamt-KPIs", style_h2))

    def kpi_cell(label, value, sub=""):
        return [
            Paragraph(label, style_kpi_lbl),
            Paragraph(value, style_kpi_val),
            Paragraph(sub,   style_kpi_sub),
        ]

    kpi_data = [[
        kpi_cell("AKTIVE KAMPAGNEN",   str(active_count),           f"{len(campaigns)} gesamt"),
        kpi_cell("GESAMTAUSGABEN",     f"{total_spend:,.0f} €",      f"Budget: {total_budget:,.0f} €"),
        kpi_cell("CONVERSIONS",        str(total_conv),             f"CPA: {cpa:,.2f} €" if cpa else "CPA: –"),
        kpi_cell("KLICKS",             f"{total_clicks:,}",         f"CTR: {ctr:.2f}%" if ctr else "CTR: –"),
        kpi_cell("IMPRESSIONEN",       f"{total_impr:,}",           ""),
    ]]
    kpi_tbl = Table(kpi_data, colWidths=[34*mm]*5)
    kpi_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), BLUE_LIGHT),
        ("BOX",           (0,0), (-1,-1), 0.5, BLUE_LIGHT),
        ("INNERGRID",     (0,0), (-1,-1), 0.5, WHITE),
        ("TOPPADDING",    (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("VALIGN",        (0,0), (-1,-1), "TOP"),
        ("ROUNDEDCORNERS",(0,0), (-1,-1), [6]),
    ]))
    story.append(kpi_tbl)
    story.append(Spacer(1, 8*mm))

    # ── KST-Tabelle ───────────────────────────────────────────
    if kst_rows:
        story.append(Paragraph("Auswertung nach Kostenstelle", style_h2))

        kst_header = [
            Paragraph("<b>Niederlassung</b>", style_center),
            Paragraph("<b>KST</b>", style_center),
            Paragraph("<b>Kampagnen</b>", style_center),
            Paragraph("<b>Ausgaben</b>", style_center),
            Paragraph("<b>Budget</b>", style_center),
            Paragraph("<b>Klicks</b>", style_center),
            Paragraph("<b>Conv.</b>", style_center),
            Paragraph("<b>CPA</b>", style_center),
        ]
        kst_tbl_data = [kst_header]
        for row in kst_rows:
            cpa_r = (row["spend"] / row["conversions"]) if row["conversions"] > 0 else None
            kst_tbl_data.append([
                row["name"],
                f"KST-{row['kostenstelle']}",
                str(row["campaign_count"]),
                f"{row['spend']:,.2f} €",
                f"{row['budget']:,.2f} €",
                f"{row['clicks']:,}",
                str(row["conversions"]),
                f"{cpa_r:,.2f} €" if cpa_r else "–",
            ])
        # Summenzeile
        kst_tbl_data.append([
            Paragraph("<b>Gesamt</b>", style_center), "",
            Paragraph(f"<b>{len(campaigns)}</b>", style_center),
            Paragraph(f"<b>{total_spend:,.2f} €</b>", style_center),
            Paragraph(f"<b>{total_budget:,.2f} €</b>", style_center),
            Paragraph(f"<b>{total_clicks:,}</b>", style_center),
            Paragraph(f"<b>{total_conv}</b>", style_center),
            Paragraph(f"<b>{cpa:,.2f} €</b>" if cpa else "<b>–</b>", style_center),
        ])

        kst_tbl = Table(kst_tbl_data,
            colWidths=[38*mm, 16*mm, 18*mm, 22*mm, 22*mm, 16*mm, 14*mm, 20*mm])
        kst_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),  (-1,0),  BLUE),
            ("TEXTCOLOR",     (0,0),  (-1,0),  WHITE),
            ("FONTNAME",      (0,0),  (-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0,0),  (-1,-1), 8.5),
            ("TOPPADDING",    (0,0),  (-1,-1), 5),
            ("BOTTOMPADDING", (0,0),  (-1,-1), 5),
            ("ROWBACKGROUNDS",(0,1),  (-1,-2), [WHITE, GRAY_LIGHT]),
            ("BACKGROUND",    (0,-1), (-1,-1), BLUE_LIGHT),
            ("FONTNAME",      (0,-1), (-1,-1), "Helvetica-Bold"),
            ("GRID",          (0,0),  (-1,-1), 0.3, colors.HexColor("#E5E7EB")),
            ("ALIGN",         (2,0),  (-1,-1), "RIGHT"),
            ("ALIGN",         (0,0),  (1,-1),  "LEFT"),
        ]))
        story.append(kst_tbl)
        story.append(Spacer(1, 8*mm))

    # ── Charts ────────────────────────────────────────────────
    if chart_paths:
        story.append(Paragraph("Grafiken", style_h2))

        for path in chart_paths:
            if os.path.exists(path):
                img = Image(path, width=165*mm, height=72*mm)
                story.append(img)
                story.append(Spacer(1, 5*mm))

    # ── Kampagnen-Detail ──────────────────────────────────────
    story.append(Paragraph("Kampagnen im Zeitraum", style_h2))

    STATUS_DE = {
        "active": "Aktiv", "pending_approval": "Warte auf Genehmigung",
        "paused": "Pausiert", "completed": "Abgeschlossen", "rejected": "Abgelehnt",
    }
    PLATFORM_DE = {"google": "Google", "microsoft": "Microsoft", "both": "Google + MS"}

    camp_header = [
        Paragraph("<b>Kampagne</b>", style_center),
        Paragraph("<b>KST</b>", style_center),
        Paragraph("<b>Plattform</b>", style_center),
        Paragraph("<b>Ausgaben</b>", style_center),
        Paragraph("<b>Conv.</b>", style_center),
        Paragraph("<b>CPA</b>", style_center),
        Paragraph("<b>Status</b>", style_center),
    ]
    camp_tbl_data = [camp_header]
    for c in campaigns[:50]:  # max 50 im PDF
        cpa_c = (c.total_cost / c.conversions) if (c.conversions and c.total_cost) else None
        camp_tbl_data.append([
            Paragraph(c.name[:48], ParagraphStyle("S", parent=styles["Normal"], fontSize=7.5)),
            f"KST-{c.kostenstelle}",
            PLATFORM_DE.get(c.platform, c.platform),
            f"{c.total_cost or 0:,.2f} €",
            str(c.conversions or 0),
            f"{cpa_c:,.2f} €" if cpa_c else "–",
            STATUS_DE.get(c.status, c.status),
        ])

    camp_tbl = Table(camp_tbl_data,
        colWidths=[55*mm, 14*mm, 20*mm, 20*mm, 13*mm, 18*mm, 30*mm])
    camp_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),  (-1,0),  BLUE),
        ("TEXTCOLOR",     (0,0),  (-1,0),  WHITE),
        ("FONTNAME",      (0,0),  (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),  (-1,-1), 7.5),
        ("TOPPADDING",    (0,0),  (-1,-1), 4),
        ("BOTTOMPADDING", (0,0),  (-1,-1), 4),
        ("ROWBACKGROUNDS",(0,1),  (-1,-1), [WHITE, GRAY_LIGHT]),
        ("GRID",          (0,0),  (-1,-1), 0.3, colors.HexColor("#E5E7EB")),
        ("ALIGN",         (3,0),  (-1,-1), "RIGHT"),
        ("ALIGN",         (0,0),  (2,-1),  "LEFT"),
    ]))
    story.append(camp_tbl)

    if len(campaigns) > 50:
        story.append(Spacer(1, 3*mm))
        story.append(Paragraph(f"... und {len(campaigns) - 50} weitere Kampagnen (nicht angezeigt).", style_body))

    # ── Footer ────────────────────────────────────────────────
    story.append(Spacer(1, 8*mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=GRAY_LIGHT))
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(
        f"Adynex · Erstellt am {datetime.now().strftime('%d.%m.%Y um %H:%M Uhr')} · adynex.de",
        style_center
    ))

    # ── Build + Send ──────────────────────────────────────────
    doc.build(story)
    output.seek(0)

    # Temp-Charts aufräumen
    for p in chart_paths:
        try:
            os.remove(p)
        except Exception:
            pass

    filename = f"Adynex_Reporting_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(
        output,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
    )
